import io
import pytz
import logging
import base64
import openpyxl

from odoo import fields, api, models
from datetime import timedelta
from odoo.exceptions import ValidationError
from odoo.tools import pycompat

_logger = logging.getLogger(__name__)

class ReportMovementHistory(models.Model):
    _name = "report.movement.history"
    _description = "Report Movement History"


    def _default_start_date(self):
        return fields.Datetime.now() - timedelta(days=1)

    start_date = fields.Datetime(string="Fecha de Inicio", required=True, 
                                default=_default_start_date,
                                help="Fecha inicial del rango.")
    end_date = fields.Datetime(string="Fecha Final", required=True, 
                                default=fields.Datetime.now,
                                help="Fecha final del rango.")
    location_id = fields.Many2one('stock.location', string="Localización", required=True,
                                help="Elegir la localización.")
    product_ids = fields.Many2many( 'product.product', string='Productos',
                                help="Elegir el producto."
    )

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date and self.end_date and self.end_date < self.start_date:
            self.end_date = self.start_date

    @api.onchange('end_date')
    def _onchange_end_date(self):
        if self.end_date and self.end_date < self.start_date:
            self.start_date = self.end_date

    def action_print_pdf(self):
        if self.start_date > self.end_date:
            raise ValidationError('La fecha de inicio debe ser menor que la fecha de finalización')
        
        data = {
            'start_date': self.start_date,
            'end_date': self.end_date,
            'location': self.location_id.id,
            'product_ids': self.product_ids.mapped('id'),
        }

        data.update(
            self.get_sale_details(
                    data['start_date'], 
                    data['end_date'], 
                    data['location'], 
                    data['product_ids']
            )
        )

        data =  {
            'type': 'ir.actions.report',
            'report_name': 'report_movement_history.report_moves_history',  # Reemplaza con el nombre de tu reporte
            'report_type': 'qweb-pdf',
            'data': {
                'form': data,  # Aquí está tu diccionario con los datos
            }, 
            'context': self.env.context,
        }
        
        return data

    @api.model
    def get_sale_details(self, start_date=False, end_date=False, location=False, products=False):
        aux_start_date = start_date
        aux_end_date = end_date

        if aux_start_date:
            aux_start_date = fields.Datetime.from_string(aux_start_date)
        else:
            # start by default today 00:00:00
            user_tz = pytz.timezone(self.env.context.get('tz') or self.env.user.tz or 'UTC')
            today = user_tz.localize(fields.Datetime.from_string(fields.Date.context_today(self)))
            aux_start_date = today.astimezone(pytz.timezone('UTC'))

        if aux_end_date:
            aux_end_date = fields.Datetime.from_string(aux_end_date)
            # avoid a aux_end_date smaller than aux_start_date
            if (aux_end_date < aux_start_date):
                aux_end_date = aux_start_date + timedelta(days=1, seconds=-1)
        else:
            # stop by default today 23:59:59
            aux_end_date = aux_start_date + timedelta(days=1, seconds=-1)

        domain = [
            '&', '&', 
            ('state', '=', 'done'),
            ('product_id', 'in', products),
            ('date', '<=', fields.Datetime.to_string(aux_end_date)),
            '|',
            ('location_id', '=', location),
            ('location_dest_id', '=', location),
        ]

        all_moves = self.env['stock.move.line'].search(domain, order='date')
        prods = self.env['product.product'].search([('id', 'in', products)])

        return{
            #'products': prods,
            'data': list(self._process_moves(products, all_moves, aux_start_date, location))
        }
        
    @api.model
    def _process_moves(self, products, all_moves, date_start, location):
        for product in products:
            lst = []
            total_out = 0
            total_in = 0
            opening_balance = 0
            balance = 0
            moves = all_moves.filtered(lambda r: r.product_id.id==product)

            if moves:
                # Cálculo del balance inicial (antes de date_start)
                balance = round(sum(
                    -line.quantity if line.location_id.id == location else line.quantity
                    for line in moves
                    if line.date <= date_start
                ), 4)
                opening_balance = balance
                moves = moves.filtered(lambda r: r.date >= date_start)
                for line in moves:
                    balance_actual_sin_modificar = "{0:.4f}".format(balance)
                    temp_dict = {}
                    
                    if location == line.location_id.id:
                        # Es una SALIDA (el producto sale de esta ubicación)
                        temp_dict['out'] = line.quantity 
                        temp_dict['in'] = 0
                        balance -= line.quantity  # Restar del balance
                        total_out += line.quantity
                    else:
                        # Es una ENTRADA (el producto entra a esta ubicación)
                        temp_dict['in'] = line.quantity 
                        temp_dict['out'] = 0
                        balance += line.quantity  # Sumar al balance
                        total_in += line.quantity
                    
                    # Agregar información adicional
                    temp_dict['date'] = line.date
                    temp_dict['picking_type'] = self.substitute(line.picking_code or '')  # Usar picking_code en lugar de picking_type_id.code
                    temp_dict['balance'] = "{0:.4f}".format(balance)
                    temp_dict['reference'] = line.reference or line.move_id.reference or line.picking_id.name or ''
                    temp_dict['balance_actual_sin_modificar'] = balance_actual_sin_modificar

                    lst.append(temp_dict)
                
            yield {
                'opening_balance': "{0:.4f}".format(opening_balance),
                'balance': "{0:.4f}".format(balance),
                'total_in': "{:.4f}".format(total_in),
                'total_out': "{:.4f}".format(total_out),
                'lst': lst,
                'product_data': product,
            }

    @staticmethod
    def substitute(code):
        return code.replace('_', " ").title() if code else code


    def action_print_csv(self):
        # Crear el contenido del CSV
        output = io.BytesIO()
        writer = pycompat.csv_writer(output, quoting=1)
        #aqui obtengo la data que deberia obtener
        if self.start_date > self.end_date:
            raise ValidationError('La fecha de inicio debe ser menor que la fecha de finalización')
        
        data = {
            'start_date': self.start_date,
            'end_date': self.end_date,
            'location': self.location_id.id,
            'product_ids': self.env['product.product'].search([('active', '=', True)]).ids,
        }

        data.update(
            self.get_sale_details(
                    data['start_date'], 
                    data['end_date'], 
                    data['location'], 
                    data['product_ids']
            )
        )
        
        # Escribir el encabezado
        writer.writerow([
            'Fecha', 'Tipo Movimiento', 'Producto', 'Referencia', 'Cantidad', 'Cantidad Contada', 'Diferencia'
        ])

        for item in data['data']:
            product_data = self.env['product.product'].browse(item['product_data'])
            full_name = product_data.name
            if product_data.default_code:
                full_name = "[" + product_data.default_code + "] " + product_data.name
            if item['lst']: 
                for move in item['lst']:
                    valor = "{:.4f}".format(move['in'] - move['out'])
                    writer.writerow([
                        move['date'],
                        move['picking_type'],
                        full_name,
                        move['reference'],
                        move['balance_actual_sin_modificar'],
                        move['balance'],
                        valor,
                    ])
            else:
                writer.writerow([
                    "-", "-", full_name, "-", "-", "-", "-"
                ])  


        # Preparar la respuesta para descargar el archivo
        output.seek(0)
        file_content = output.getvalue()
        
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/?model=ir.attachment&field=datas&filename_field=name&id=%s' % self._create_attachment(file_content).id,
            'target': 'self',
        }

    def _create_attachment(self, file_content):
        # Crear un attachment con el contenido del CSV
        attachment = self.env['ir.attachment'].create({
            'name': 'Reporte de Historial de Movimiento.csv',
            'datas': base64.b64encode(file_content),
            'type': 'binary',
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'text/csv',
        })
        return attachment


    def action_print_xlsx(self):
        if self.start_date > self.end_date:
            raise ValidationError('La fecha de inicio debe ser menor que la fecha de finalización')
        
        # Crear un nuevo libro de Excel
        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Reporte de Movimientos"

        # Escribir el encabezado
        headers = ['Fecha', 'Tipo Movimiento', 'Producto', 'Referencia', 'Unidad de Medida', 'Costo del Producto', 'Cantidad', 'Cantidad Contada', 'Diferencia']
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        # Obtener los datos
        data = {
            'start_date': self.start_date,
            'end_date': self.end_date,
            'location': self.location_id.id,
            'product_ids': self.env['product.product'].search([('active', '=', True)]).ids,
        }

        data.update(
            self.get_sale_details(
                data['start_date'], 
                data['end_date'], 
                data['location'], 
                data['product_ids']
            )
        )

        # Escribir los datos
        row = 2
        for item in data['data']:
            product_data = self.env['product.product'].browse(item['product_data'])
            full_name = product_data.name
            product_uom = product_data.uom_name
            product_standard_price = "{:.4f}".format(product_data.standard_price)
            if product_data.default_code:
                full_name = f"[{product_data.default_code}] {product_data.name}"
            
            if item['lst']: 
                for move in item['lst']:
                    valor = "{:.4f}".format(move['in'] - move['out'])
                    worksheet.cell(row=row, column=1, value=move['date'])
                    worksheet.cell(row=row, column=2, value=move['picking_type'])
                    worksheet.cell(row=row, column=3, value=full_name)
                    worksheet.cell(row=row, column=4, value=move['reference'])
                    worksheet.cell(row=row, column=5, value=product_uom)
                    worksheet.cell(row=row, column=6, value=product_standard_price)
                    worksheet.cell(row=row, column=7, value=move['balance_actual_sin_modificar'])
                    worksheet.cell(row=row, column=8, value=move['balance'])
                    worksheet.cell(row=row, column=9, value=valor)
                    row += 1
            else:
                worksheet.cell(row=row, column=3, value=full_name)
                row += 1

        # Guardar el libro de trabajo en el buffer
        workbook.save(output)
        output.seek(0)
        file_content = output.getvalue()

        # Crear y devolver el attachment
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/?model=ir.attachment&field=datas&filename_field=name&id=%s' % self._create_xlsx_attachment(file_content).id,
            'target': 'self',
        }

    def _create_xlsx_attachment(self, file_content):
        attachment = self.env['ir.attachment'].create({
            'name': 'Reporte de Historial de Movimiento.xlsx',
            'datas': base64.b64encode(file_content),
            'type': 'binary',
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return attachment